import logging
import traceback
import requests
from bs4 import BeautifulSoup
from copy import copy
from datetime import date as datetimedate, timedelta
from shutil import copyfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator


class WorkbookManager:

    """
    Reads and writes to excel spreadsheet for fundsbook.

    The workbook manager class contains data after reading:
        - missing_funds: the missing funds for the funds sheet
        - ranking_ids: the ranking ids that need to be scraped for the ranking sheet
            (or top sheets)
    """

    def __init__(
        self, path: str = "FundsBook.xlsm", backup: bool = True
    ):  # str='FundsBook.xlsx'
        self.path = path
        self.backup = backup
        # self.wb = load_workbook(path, read_only=False, keep_vba=True)
        self.missing_funds = None
        self.ranking_ids = None

        # if self.backup:
        # copyfile(path, f"app/workbooks/backup_{datetimedate.today()}_{path}")

    @staticmethod
    def date_by_adding_business_days(from_date, add_days):
        business_days_to_add = add_days
        current_date = from_date
        while business_days_to_add > 0:
            current_date += timedelta(days=1)
            weekday = current_date.weekday()
            if weekday >= 5:  # sunday = 6
                continue
            business_days_to_add -= 1
        return current_date

    def read_funds(
        self, sheet: str = "基金日记", progress_callback=None, progress_callback_num=None
    ) -> bool:

        """
        Reads and saves the missing cells that need to be filled in in the 'Funds' sheet.
        Returns true on success, false on failure.

        The missing_funds data is populated in the following way.
        It is a dictionary where each item has the format:
            <id>: {
                'column': <column_no>,
                'missing_dates': <list_dates>
            }

        Where
            <id> is a string of the fund id;

            <column_no> is an int of the column number where
                this company is in the spreadsheet;

            <list_dates> is an array of two-tuples (<row_no>, <date>),
                where the first element is the row number for that cell
                and the second element is the date corresponding to that cell.

        Each item of missing_funds represents a company fund in the funds sheet.
        Elements in 'missing_dates' means it found cells that had a date but no
        price.
        """

        try:
            logging.info("(Read funds) Finding missing data for funds sheet")
            progress_callback.emit(f"(Read funds) Finding missing data for funds sheet")
            ws = self.wb[sheet]
            price_columns = {}

            """
             NOTE by jj: this part of auto fill-in dates han been done in Excel itself with more functions such as excluding local holidays and weekends
           
            # Move all price history down and add dates until we reach TODAY's date
            # Find empty box to start putting dates first:
            i = 10
            empty_date = ws.cell(row=i, column=11).value
            while empty_date is None:
                i += 1
                empty_date = ws.cell(row=i, column=11).value

            # empty_date -> is acutally the cell containing the last inserted date!!
            
            # grab today's date:
            today = datetimedate.today()
            
            # insert until today:
            current_day = self.date_by_adding_business_days(empty_date, 1)
            
            
            move_down = (today-empty_date.date()).days
            move_down = move_down - ((move_down//7)*2)
                        
            ws.move_range(f'K9:{get_column_letter(ws.max_column)}{ws.max_row}', rows=move_down, translate=True)

            for row in range(9, 9+move_down):
                ws.cell(row=row, column=11).value = None
                for col in range(2, ws.max_column+1):
                    ws.cell(row=row, column=col).value = Translator(
                        ws.cell(row=row+move_down, column=col).value,
                        f'{get_column_letter(col)}{row}'
                        ).translate_formula(row_delta=-move_down, col_delta=0)
                    if ws.cell(row=row+move_down, column=col).has_style:
                        ws.cell(row=row, column=col)._style = copy(ws.cell(row=row+move_down, column=col)._style)
                    if (col-9)%8==0:
                        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF2CC")

            row_count = 1
            while (current_day.date() <= today):
                ws.cell(row=i+move_down-row_count, column=11).value = current_day
                row_count += 1
                current_day = self.date_by_adding_business_days(current_day, 1)
            """

            # Get the columns of all company funds ids
            for i in range(ws.max_column + 1):
                if ws.cell(row=1, column=i + 1).value is not None:
                    price_columns[ws.cell(row=1, column=i + 1).value] = i + 1

            # Construct the missing_funds data
            self.missing_funds = self._get_missing_price_funds(ws, price_columns)

        except Exception as e:
            logging.error(f"(Read funds) Failed to read missing data for funds sheet")
            progress_callback.emit(
                f"(Read funds) Failed to read missing for funds sheet"
            )
            logging.error(traceback.format_exc())
            return False

        logging.info(
            f"(Read funds) Found {len(self.missing_funds)} that have missing dates to fill"
        )
        progress_callback.emit(
            f"(Read funds) Found {len(self.missing_funds)} that have missing dates to fill"
        )
        return True

    def _get_missing_price_funds(self, ws, price_columns) -> dict:
        """
        A helper function to parse the missing cells for the funds sheet.
        """
        row = 9  # pass the first few rows that may have contents other than price
        current = None
        missing_price_funds = {}

        for id, col in price_columns.items():
            # Skip the blank template
            if id == "000000":
                continue

            # get the last blank row in this column
            current_row = row
            while current is None or current == "":
                current = ws.cell(row=current_row, column=col).value
                # print(current)
                current_row += 1

            # print(current_row)
            missing_price_funds[id] = {"column": col, "missing-dates": []}

            current_row -= 1

            # record all the missing dates
            """ column=1 is the date column """
            while ws.cell(row=current_row, column=1).value is not None:
                # print(current_row)
                missing_price_funds[id]["missing-dates"].append(
                    (
                        current_row,
                        (ws.cell(row=current_row, column=1).value).strftime("%Y-%m-%d"),
                    )
                )
                current_row -= 1

            # reset and repeat
            current = None

        logging.info(
            f"(Read missing funds) Done reading missing cells for funds sheet {missing_price_funds}"
        )
        return missing_price_funds

    def write_funds(
        self,
        funds_data: dict = {},
        sheet: str = "基金日记",
        progress_callback=None,
        progress_callback_num=None,
    ) -> bool:
        """
        Writes the missing funds to the workbook with the given data.
        The missing_funds should have been initialized beforehand.
        Returns true on success, false on failure.

        funds_data is a dictionary containing items with the format:
        <id> : {<date> : <value>, ...}
        """
        try:
            ws = self.wb[sheet]
            count = 0

            for id, dates in self.missing_funds.items():
                col = dates["column"]
                for date in dates["missing-dates"]:
                    row = date[0]
                    missing_date = date[1]
                    try:
                        try:
                            ws.cell(row=row, column=col).value = float(
                                funds_data[str(id)][missing_date]
                            )
                        except ValueError:
                            ws.cell(row=row, column=col).value = funds_data[str(id)][
                                missing_date
                            ]
                        count += 1
                    except KeyError as e:
                        logging.warning(
                            f"(Write funds) Tried to write to ({row},{col}) but did not find id or missing date in funds data (KeyError {e})"
                        )
                        progress_callback.emit(
                            f"------------------------- (Write funds) Tried to write to ({row},{col}) but did not find id or missing date in funds data (KeyError {e})"
                        )

            logging.info(f"(Write funds) Done writing {count} new cells")
            progress_callback.emit(f"(Write funds) Done writing {count} cells")

            return True

        except Exception as e:
            logging.error("(Write funds) Failed to write funds data")
            progress_callback.emit(f"(Write funds) Failed to write funds data")
            logging.error(traceback.format_exc())
            return False

    def read_rankings(
        self, sheet: str = "基金排队", progress_callback=None, progress_callback_num=None
    ) -> bool:
        """
        Reads and saves the ids of companies in the 'Rankings' sheet.
        Only non-empty id values are considered. Returns true on success, false on failure.

        The ranking_ids data is populated in the following way:
        It is a list of openpyxl Cell objects. Each Cell object represents
        where the company id was written. The value can be accessed through Cell.value
        """
        try:
            ws = self.wb[sheet]
            self.ranking_ids = [
                id
                for id in ws["D"]
                if not str(id.value).startswith("=") and id.value is not None
            ]

            logging.info(f"(Read rankings) Retrieved {len(self.ranking_ids)} ids")
            progress_callback.emit(
                f"(Read rankings) Retrieved {len(self.ranking_ids)} ids"
            )
            logging.debug(
                f"[read rankings] Retrieved all ranking ids {[str(a.value) for a in self.ranking_ids]}"
            )
            return True

        except Exception as e:
            logging.error(f"(Read rankings) Failed to retrieve ranking ids.")
            progress_callback.emit(f"(Read rankings) Failed to retrieve ranking ids.")
            logging.error(traceback.format_exc())
            return False

    def write_rankings(
        self,
        ranking_data: dict = {},
        sheet: str = "基金排队",
        ids_override: list = None,
        progress_callback=None,
        progress_callback_num=None,
    ) -> bool:
        """
        Writes missing ranking values to the sheet with the given data.

        The ranking_data is a dictionary with the following format:
        <id> : [<value>, ...]

        The ids_override is a list of two-tuples with the following format:
        (<id>, <name>)

        The ranking data is written to the ranking sheet one by one.

        If the ids_override is passed a list of ids, the function writes the
        rows based on the order given in ids_override ids. The data is still required
        to be in ranking_data.

        The ids_override parameter is meant to be used to write to top sheet rankings.
        """

        ws = self.wb[sheet]
        id_col = 4  # fund's code
        low_col = 6  # date
        high_col = 17  # 5 year value
        count = 0

        try:
            if not ids_override:
                for i in range(ws.max_row):
                    id = ws.cell(row=i + 1, column=id_col).value
                    if not str(id).startswith("=") and id is not None:
                        for idx, c in enumerate(range(low_col, high_col + 1)):
                            try:
                                if idx == 0:

                                    # column D of sheet 基金排队, this is the found's code
                                    ws.cell(row=i + 1, column=c).value = ranking_data[
                                        str(id)
                                    ][idx]

                                    """ jj added the following 3 lines, still need to figure out which are date and price columns """
                                    # column E of sheet 基金排队, fund's name
                                    ws.cell(
                                        row=i + 1, column=id_col + 1
                                    ).value = self._get_name_from_id(id)
                                    ws.cell(
                                        row=i + 1, column=id_col + 1
                                    ).value = f'=HYPERLINK("https://fund.eastmoney.com/{id}.html", "{self._get_name_from_id(id)}")'
                                    ws.cell(row=i + 1, column=id_col + 1).font = Font(
                                        size=9, color="0000ff"
                                    )

                                    """  
                                    # columns R,S,T of sheet 基金排队
                                    ws.cell(row=i+1, column=id_col+14).value = 'R column'
                                    ws.cell(row=i+1, column=id_col+15).value = 'S column'
                                    ws.cell(row=i+1, column=id_col+16).value = 'T column'
                                    """

                                    """
                                    ??? where are the code lines to write column F (Date column of sheet 基金排队) ???
                                    """

                                elif ranking_data[str(id)][idx].endswith("%"):
                                    percent = ranking_data[str(id)][idx][:-1]
                                    percent = float(percent) / 100
                                    ws.cell(row=i + 1, column=c).style = "Percent"
                                    ws.cell(row=i + 1, column=c).number_format = "0.00%"
                                    ws.cell(row=i + 1, column=c).value = percent

                                else:
                                    try:
                                        # column G of sheet 基金排队, the Price column
                                        ws.cell(row=i + 1, column=c).value = float(
                                            ranking_data[str(id)][idx]
                                        )

                                    except ValueError:
                                        ws.cell(
                                            row=i + 1, column=c
                                        ).value = ranking_data[str(id)][idx]

                                count += 1
                            except KeyError as e:
                                logging.warning(
                                    f"(Write ranking) Tried to write to ({i+1, c}) but data was missing (KeyError {e}) for {id}"
                                )
                                progress_callback.emit(
                                    f"------------------------- (Write ranking) Tried to write to ({i+1, c}) but data was missing (KeyError {e}) for {id}"
                                )

                            except IndexError as e:
                                logging.warning(
                                    f"(Write ranking) Tried to write to ({i+1, c}) but data was missing (IndexError {e}) for {id}"
                                )
                                progress_callback.emit(
                                    f"------------------------- (Write ranking) Tried to write to ({i+1, c}) but data was missing (IndexError {e}) for {id}"
                                )

            else:
                # Fix for writing to topx sheets
                # Add to this list to not write to rows
                exclude_rows = [1]
                old_ids = []
                i = 0

                # move current sheet down
                self.copy_top_range(
                    sheet=sheet,
                    progress_callback=progress_callback,
                    progress_callback_num=progress_callback_num,
                )

                while len(ids_override) > 0:
                    if (i + 1) not in exclude_rows:
                        id = ws.cell(row=i + 1, column=id_col).value

                        if not str(id).startswith("="):
                            a = ids_override.pop(0)
                            new_id = a[0]
                            new_name = a[1]
                            ws.cell(row=i + 1, column=id_col).value = new_id

                            # Write the link to the company name
                            ws.cell(
                                row=i + 1, column=id_col + 1
                            ).value = f'=HYPERLINK("https://fund.eastmoney.com/{new_id}.html", "{new_name}")'

                            old_ids.append(a)

                            for idx, c in enumerate(range(low_col, high_col + 1)):
                                try:
                                    if idx == 0:
                                        ws.cell(
                                            row=i + 1, column=c
                                        ).value = ranking_data[str(new_id)][idx]
                                    elif ranking_data[str(new_id)][idx].endswith("%"):
                                        percent = ranking_data[str(new_id)][idx][:-1]
                                        percent = float(percent) / 100
                                        ws.cell(row=i + 1, column=c).style = "Percent"
                                        ws.cell(
                                            row=i + 1, column=c
                                        ).number_format = "0.00%"
                                        ws.cell(row=i + 1, column=c).value = percent
                                    else:
                                        try:
                                            ws.cell(row=i + 1, column=c).value = float(
                                                ranking_data[str(new_id)][idx].strip()
                                            )
                                            ws.cell(
                                                row=i + 1, column=c
                                            ).number_format = "0.0000"
                                        except ValueError:
                                            ws.cell(
                                                row=i + 1, column=c
                                            ).value = ranking_data[str(new_id)][idx]

                                    count += 1
                                except KeyError as e:
                                    logging.error(
                                        f"(Write ranking) Tried to write to ({i+1, c}) but data was missing (KeyError {e}) for {id}"
                                    )
                                    progress_callback.emit(
                                        f"------------------------- (Write ranking) Tried to write to ({i+1, c}) but data was missing (KeyError {e}) for {id}"
                                    )

                                except IndexError as e:
                                    logging.error(
                                        f"(Write ranking) Tried to write to ({i+1, c}) but data was missing (IndexError {e}) for {id}"
                                    )
                                    progress_callback.emit(
                                        f"------------------------- (Write ranking) Tried to write to ({i+1, c}) but data was missing (IndexError {e}) for {id}"
                                    )

                    i += 1

                ids_override = old_ids

            logging.info(f"(Write ranking) Done writing {count} new cells")
            progress_callback.emit(f"(Write ranking) Done writing {count} new cells")
            return True
        except Exception as e:
            logging.error(f"(Write ranking) Failed to write ranking data")
            progress_callback.emit(f"(Write ranking) Failed to write ranking data")
            logging.error(traceback.format_exc())
            return False

    def copy_top_range(
        self, sheet: str = "top50混合", progress_callback=None, progress_callback_num=None
    ) -> bool:
        try:
            ws = self.wb[sheet]
            id_col = 4  # fund's code, the column for #1 to #50 (company ids)

            # move the whole page down {row_down} rows, except the first row that's the title row
            row_down = 52

            """ if want to write over the old data(no move down 52 rows), just comment out the following blocks """

            # before move dowm 52 rows, clear columns A and R to AA
            for idx, row_i in enumerate(range(2, 52), 1):
                # column A：持仓/清仓情况
                ws.cell(row=row_i, column=id_col - 3).value = ""
                # column R to AA
                ws.cell(row=row_i, column=id_col + 14).value = ""
                ws.cell(row=row_i, column=id_col + 15).value = ""
                ws.cell(row=row_i, column=id_col + 16).value = ""
                ws.cell(row=row_i, column=id_col + 17).value = ""
                ws.cell(row=row_i, column=id_col + 18).value = ""
                ws.cell(row=row_i, column=id_col + 19).value = ""
                ws.cell(row=row_i, column=id_col + 20).value = ""
                ws.cell(row=row_i, column=id_col + 21).value = ""
                ws.cell(row=row_i, column=id_col + 22).value = ""
                ws.cell(row=row_i, column=id_col + 23).value = ""
                # ws.cell(row=row_i, column=id_col+24).value = "" NOTE: we need this to calculate position difference

            ws.move_range(f"A2:AB{ws.max_row+1}", rows=row_down, translate=True)
            # copy the first row (title row) down
            for c in range(1, ws.max_column - 3):
                ws.cell(row=row_down + 1, column=c).value = ws.cell(
                    row=1, column=c
                ).value
                ws.cell(row=row_down + 1, column=c).fill = PatternFill(
                    "solid", fgColor="FCE4D6"
                )
                # ws.cell(row=row_down+1, column=c).font = Font(name='宋体', size=9, bold=True)
                ws.cell(row=row_down + 1, column=c).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            """ if want to write over the old data(no move down 52 rows), just comment out the above blocks """

            # adding numbers, formatting and styles to the new opened up space
            # this will go through all 50 rows no matter how many funds data need to write (maximum 50)
            for idx, row_i in enumerate(range(2, 52), 1):

                # column A：持仓/清仓情况
                ws.cell(
                    row=row_i, column=id_col - 3
                ).value = f'=IFERROR(IF(MATCH(D{row_i},基金日记!$1:$1,FALSE),"持仓","error1"),IFERROR(IF(MATCH(AB{row_i},基金日记!$1:$1,FALSE),"持仓","error2"),IFERROR(IF(MATCH(D{row_i},历史日记!$1:$1,FALSE),"X","error3"),IFERROR(IF(MATCH(AB{row_i},历史日记!$1:$1,FALSE),"X","error4"),"N"))))'
                ws.cell(row=row_i, column=id_col - 3).font = Font(size=9)

                # column B：排名上升/下降情况
                ws.cell(
                    row=row_i, column=id_col - 2
                ).value = f'=IFERROR((INDEX(C54:C103,MATCH(D{row_i},D54:D103,FALSE),1)-C{row_i}),"new")'
                ws.cell(row=row_i, column=id_col - 2).alignment = Alignment(
                    horizontal="right", vertical="center"
                )

                # column C: fund's ranking position
                ws.cell(row=row_i, column=id_col - 1).value = idx
                ws.cell(row=row_i, column=id_col - 1).fill = PatternFill(
                    "solid", fgColor="FCE4D6"
                )

                # column D: fund's code
                ws.cell(row=row_i, column=id_col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws.cell(row=row_i, column=id_col).fill = PatternFill(
                    "solid", fgColor="D2E2FF"
                )

                # column E: fund's name
                ws.cell(row=row_i, column=id_col + 1).font = Font(size=9)

                # column F: date
                ws.cell(row=row_i, column=id_col + 2).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws.cell(row=row_i, column=id_col + 2).fill = PatternFill(
                    "solid", fgColor="D2E2FF"
                )

                # column G: price
                ws.cell(row=row_i, column=id_col + 3).fill = PatternFill(
                    "solid", fgColor="D2E2FF"
                )

                """ column R：持仓成本 """
                # ws.cell(row=row_i, column=id_col+14).value = f'=IFERROR(INDEX(基金排队!D$2:AA$200, MATCH(D{row_i},基金排队!D$2:D$200,0),15),"")'
                ws.cell(
                    row=row_i, column=id_col + 14
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),15),IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!AB$3:AB$200,0),15),IFERROR(IF(MATCH(D{row_i},历史排队!D$2:D$200,0),"已清仓"),IFERROR(IF(MATCH(D{row_i},历史排队!AB$2:AB$200,0),"已清仓"),"从未建仓"))))'
                ws.cell(row=row_i, column=id_col + 14).fill = PatternFill(
                    "solid", fgColor="D2E2FF"
                )
                ws.cell(row=row_i, column=id_col + 14).number_format = "#,##0.0000￥"
                ws.cell(row=row_i, column=id_col + 14).font = Font(size=9)

                """ column S：持仓收益 """
                # ws.cell(row=row_i, column=id_col+15).value = f'=IFERROR(INDEX(基金排队!D$2:AA$200, MATCH(D{row_i},基金排队!D$2:D$200,0),16),"")'
                ws.cell(
                    row=row_i, column=id_col + 15
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),16),IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!AB$3:AB$200,0),16),IFERROR(IF(MATCH(D{row_i},历史排队!D$2:D$200,0),"已清仓"),IFERROR(IF(MATCH(D{row_i},历史排队!AB$2:AB$200,0),"已清仓"),"从未建仓"))))'
                ws.cell(row=row_i, column=id_col + 15).fill = PatternFill(
                    "solid", fgColor="D2E2FF"
                )
                ws.cell(row=row_i, column=id_col + 15).number_format = "#,##0.0000￥"
                ws.cell(row=row_i, column=id_col + 15).font = Font(size=9)

                """ column T：盈利百分比 % """
                # ws.cell(row=row_i, column=id_col+16).value = f'=IFERROR(INDEX(基金排队!D$2:AA$200, MATCH(D{row_i},基金排队!D$2:D$200,0),17),"")'
                ws.cell(
                    row=row_i, column=id_col + 16
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),17),IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!AB$3:AB$200,0),17),IFERROR(IF(MATCH(D{row_i},历史排队!D$2:D$200,0),"已清仓"),IFERROR(IF(MATCH(D{row_i},历史排队!AB$2:AB$200,0),"已清仓"),"从未建仓"))))'
                ws.cell(row=row_i, column=id_col + 16).fill = PatternFill(
                    "solid", fgColor="D2E2FF"
                )
                ws.cell(row=row_i, column=id_col + 16).number_format = "0.00%"
                ws.cell(row=row_i, column=id_col + 16).font = Font(size=9)

                ws.cell(
                    row=row_i, column=id_col + 17
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),18),"")'  # U column：夏普比

                ws.cell(
                    row=row_i, column=id_col + 18
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),19),"")'  # V column：波动率（标准差）
                ws.cell(
                    row=row_i, column=id_col + 18
                ).number_format = "0.00%"  # V column

                ws.cell(
                    row=row_i, column=id_col + 19
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),20),"")'  # W column：最大回撤
                ws.cell(
                    row=row_i, column=id_col + 19
                ).number_format = "0.00%"  # W column

                ws.cell(
                    row=row_i, column=id_col + 20
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),21),"")'  # X column：收益回撤比

                ws.cell(
                    row=row_i, column=id_col + 21
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),22),"")'  # Y column：基金成立日期
                ws.cell(
                    row=row_i, column=id_col + 21
                ).number_format = "YYYY-MM-DD"  # = 2021-10-03, if using MMM, that will be 2021-Oct-03. can be 'YYYY MM DD' too

                ws.cell(
                    row=row_i, column=id_col + 22
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),23),"")'  # Z column：基金经理姓名
                ws.cell(row=row_i, column=id_col + 22).fill = PatternFill(
                    "solid", fgColor="D2E2FF"
                )
                ws.cell(row=row_i, column=id_col + 22).font = Font(size=9)

                ws.cell(
                    row=row_i, column=id_col + 23
                ).value = f'=IFERROR(INDEX(基金排队!D$3:AA$200, MATCH(D{row_i},基金排队!D$3:D$200,0),24),"")'  # AA column：基金规模（亿元）
                ws.cell(row=row_i, column=id_col + 23).fill = PatternFill(
                    "solid", fgColor="D2E2FF"
                )

                # AB column：C type fund code
                ws.cell(
                    row=row_i, column=id_col + 24
                ).value = f'=IFERROR(IF(INDEX(基金排队!$D$3:$AB$200, MATCH($D{row_i},基金排队!$D$3:$D$200,0),25)="","",IFERROR(IF(MATCH(INDEX(基金排队!$D$3:$AB$200, MATCH($D{row_i},基金排队!$D$3:$D$200,0),25),$D$2:$D$51,0),"","error"),INDEX(基金排队!$D$3:$AB$200, MATCH($D{row_i},基金排队!$D$3:$D$200,0),25))),"")'

            logging.info(
                f"(Copy top range) Moved top sheet down {row_down} rows and created new formatting on top"
            )
            progress_callback.emit(
                f"(Copy top range) Moved top sheet down {row_down} rows and created new formatting on top"
            )
            return True

        except Exception as e:
            logging.error("(Copy top range) Failed to copy top range")
            progress_callback.emit(f"(Copy top range) Failed to copy top range")
            logging.error(traceback.format_exc())
            return False

    def buy_funds(
        self, id, amount, date, sheet: str = "基金日记", progress_callback=None
    ) -> bool:
        try:
            ws = self.wb[sheet]
            new = False

            # Find the column for the company id.
            col = 1
            current_col_val = ws.cell(row=1, column=col).value
            while current_col_val != id:
                col += 1
                current_col_val = ws.cell(row=1, column=col).value

                if col >= ws.max_column:
                    logging.warning(
                        f"(buy funds) did not find id {id} in the worksheet. Creating new column"
                    )
                    progress_callback.emit(
                        f"(buy funds) did not find id {id} in the worksheet. Creating new column"
                    )
                    col = self._create_new_funds_column(id=id, open_sheet=ws)
                    new = True
                    break

            # Find the row for the date.
            row = 9
            current_row_val = ws.cell(row=row, column=11).value
            while (
                current_row_val is None or current_row_val.strftime("%Y-%m-%d") != date
            ):
                row += 1
                current_row_val = ws.cell(row=row, column=11).value

                if row >= ws.max_row:
                    logging.error(
                        f"(buy funds) did not find date {date} in the worksheet."
                    )
                    progress_callback.emit(
                        f"(buy funds) did not find date {date} in the worksheet."
                    )
                    return False

            # Insert the amount.
            ws.cell(row=row, column=col - 3).value = float(amount)
            if new:
                ws.cell(row=row + 1, column=col).value = float(1.0)

        except:
            logging.error(f"(buy funds) Failed to buy funds.")
            progress_callback.emit(f"(buy funds) Failed to buy funds.")
            logging.error(traceback.format_exc())
            progress_callback.emit(traceback.format_exc())
            return False

        logging.info(
            f"(buy funds) Inserted {amount} for {id} at date {date}. WAIT FOR THE THREAD TO FINISH FIRST."
        )
        progress_callback.emit(f"(buy funds) Inserted {amount} for {id} at date {date}")
        return True

    def _create_new_funds_column(self, id, open_sheet) -> int:
        ws = open_sheet
        template_start_col = ws.max_column - 7
        template_end_col = ws.max_column

        for row_num in range(1, ws.max_row + 1):
            for column_num in range(template_start_col, template_end_col + 1):
                print(ws.cell(row=row_num, column=column_num).value)
                ws.cell(row=row_num, column=column_num + 8).value = Translator(
                    ws.cell(row=row_num, column=column_num).value,
                    f"{get_column_letter(column_num)}{row_num}",
                ).translate_formula(row_delta=0, col_delta=8)

                if ws.cell(row=row_num, column=column_num).has_style:
                    ws.cell(row=row_num, column=column_num + 8)._style = copy(
                        ws.cell(row=row_num, column=column_num)._style
                    )

                if column_num in [template_end_col]:
                    ws.column_dimensions[f"{get_column_letter(column_num+8)}"].width = 2
                    ws.cell(row=row_num, column=column_num + 8).fill = PatternFill(
                        "solid", fgColor="FFF2CC"
                    )

        ws.cell(row=1, column=template_end_col - 4).value = id

        ws.cell(
            row=3, column=template_end_col - 5
        ).value = f'=HYPERLINK("https://fund.eastmoney.com/{id}.html", "{self._get_name_from_id(id)}")'

        return template_end_col - 4

    def _get_name_from_id(self, id):
        try:
            r = requests.get(f"https://fund.eastmoney.com/{id}.html")
            soup = BeautifulSoup(r.content, "html.parser")
            title = soup.find("title")
            return title.text.strip()[: (title.text.strip().find("("))]
        except:
            return "Not found"

    def close(self):
        """Closes the workbook and saves it under the same name."""
        self.wb.save(self.path)
        self.wb.close()
